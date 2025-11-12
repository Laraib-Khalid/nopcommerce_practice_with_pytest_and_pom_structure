import openpyxl

class ExcelUtils:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = self.wb[sheet_name]

    def get_row_count(self):
        """Return total number of rows"""
        return self.sheet.max_row

    def get_column_count(self):
        """Return total number of columns"""
        return self.sheet.max_column

    def read_data(self, row, column):
        """Read data from a specific cell"""
        return self.sheet.cell(row=row, column=column).value

    def write_data(self, row, column, value):
        """Write data into a specific cell"""
        self.sheet.cell(row=row, column=column).value = value
        self.wb.save(self.file_path)
